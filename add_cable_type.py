"""
Adds a "Type" column (Cu/Al dropdown) after each "Cable mm²" column
in every sheet of the transformer workbook.

Steps per sheet:
  1. Find Cable-mm² columns.
  2. Snapshot all merged ranges and all formula-cell locations.
  3. Remove all merges and blank-out formula cells (so insert_cols
     cannot partially—and incorrectly—update them).
  4. Insert Type columns right-to-left.
  5. Re-apply merged ranges at their shifted positions.
  6. Add new Type-header merges (rows 2-3).
  7. Re-write all formulas at their shifted locations with column
     references updated via the same shift math.
  8. Attach Cu/Al dropdown validation.
"""

import re
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.worksheet.datavalidation import DataValidation
from copy import copy

INPUT_FILE  = 'transformers_20260613 (7).xlsx'
OUTPUT_FILE = 'transformers_20260613_updated.xlsx'


# ── helpers ───────────────────────────────────────────────────────────────

def copy_cell_style(src, dst):
    if src.has_style:
        dst.font        = copy(src.font)
        dst.fill        = copy(src.fill)
        dst.alignment   = copy(src.alignment)
        dst.border      = copy(src.border)
        dst.number_format = src.number_format


def col_shift(col_idx, cable_cols):
    """
    After inserting one extra column immediately after each cable_col,
    an original column at `col_idx` ends up at:
        col_idx + number-of-cable_cols strictly less than col_idx
    """
    return col_idx + sum(1 for c in cable_cols if c < col_idx)


_CELL_REF = re.compile(r'(\$?)([A-Z]+)(\$?\d+)')

def shift_formula(formula, cable_cols):
    """Rewrite every cell-column reference in a formula string."""
    def _replace(m):
        dollar_c, col_letters, row_part = m.group(1), m.group(2), m.group(3)
        new_col = get_column_letter(
            col_shift(column_index_from_string(col_letters), cable_cols)
        )
        return f"{dollar_c}{new_col}{row_part}"
    return _CELL_REF.sub(_replace, formula)


# ── main ──────────────────────────────────────────────────────────────────

wb = load_workbook(INPUT_FILE)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]

    # 1. Find Cable-mm² columns in row 2
    cable_cols = sorted(
        col for col in range(1, ws.max_column + 1)
        if 'Cable' in str(ws.cell(row=2, column=col).value or '')
    )
    if not cable_cols:
        print(f"  Skipping '{sheet_name}' - no Cable mm2 headers")
        continue

    print(f"Processing '{sheet_name}': "
          f"Cable mm2 at {[get_column_letter(c) for c in cable_cols]}")

    # 2a. Snapshot merged ranges
    merge_snapshot = [
        (r.min_row, r.min_col, r.max_row, r.max_col)
        for r in list(ws.merged_cells.ranges)
    ]

    # 2b. Snapshot formula cells (row, col, formula_string)
    formula_snapshot = []
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith('='):
                formula_snapshot.append((cell.row, cell.column, cell.value))

    # 2c. Snapshot header styles (for Type column styling)
    # key = cable_col → styles of the adjacent R/Y/B/N header (row 2)
    hdr_styles = {}
    for cc in cable_cols:
        src = ws.cell(row=2, column=cc + 1)
        hdr_styles[cc] = (
            copy(src.font)       if src.has_style else None,
            copy(src.fill)       if src.has_style else None,
            copy(src.alignment)  if src.has_style else None,
            copy(src.border)     if src.has_style else None,
            src.number_format,
        )
    # key = cable_col → styles of a data row cell in the cable column
    data_styles = {}
    for cc in cable_cols:
        src = ws.cell(row=4, column=cc)
        data_styles[cc] = (
            copy(src.font)       if src.has_style else None,
            copy(src.fill)       if src.has_style else None,
            copy(src.alignment)  if src.has_style else None,
            copy(src.border)     if src.has_style else None,
            src.number_format,
        )

    # 3. Remove all merges; blank out formula cells so insert_cols
    #    cannot incorrectly touch them
    for (r1, c1, r2, c2) in merge_snapshot:
        ws.unmerge_cells(start_row=r1, start_column=c1,
                         end_row=r2,   end_column=c2)
    for (r, c, _) in formula_snapshot:
        ws.cell(row=r, column=c).value = None

    # 4. Insert Type columns right-to-left
    for cc in reversed(cable_cols):
        ws.insert_cols(cc + 1)

    # 5. Re-apply all merges at shifted positions
    for (r1, c1, r2, c2) in merge_snapshot:
        new_c1 = col_shift(c1, cable_cols)
        new_c2 = col_shift(c2, cable_cols)
        ws.merge_cells(start_row=r1, start_column=new_c1,
                       end_row=r2,   end_column=new_c2)

    # 6. Write Type headers, add Type merges, style everything
    for i, cc in enumerate(cable_cols, start=1):
        # Final column index of this Type column after all insertions
        type_col = cc + i

        hdr = ws.cell(row=2, column=type_col)
        hdr.value = "Type"
        font, fill, align, border, nf = hdr_styles[cc]
        if font:   hdr.font      = font
        if fill:   hdr.fill      = fill
        if align:  hdr.alignment = align
        if border: hdr.border    = border
        hdr.number_format = nf

        ws.merge_cells(start_row=2, start_column=type_col,
                       end_row=3,   end_column=type_col)

        # Style data rows
        font, fill, align, border, nf = data_styles[cc]
        for row in range(4, ws.max_row + 1):
            cell = ws.cell(row=row, column=type_col)
            if font:   cell.font      = font
            if fill:   cell.fill      = fill
            if align:  cell.alignment = align
            if border: cell.border    = border
            cell.number_format = nf

        # Cu/Al dropdown
        col_ltr = get_column_letter(type_col)
        dv = DataValidation(
            type="list", formula1='"Cu,Al"',
            allow_blank=True, showDropDown=False,
        )
        dv.error       = "Please choose Cu or Al"
        dv.errorTitle  = "Invalid value"
        dv.prompt      = "Choose cable material"
        dv.promptTitle = "Cable Type"
        ws.add_data_validation(dv)
        dv.sqref = f"{col_ltr}4:{col_ltr}{ws.max_row}"

    # 7. Re-insert formulas at their shifted positions with updated refs
    for (r, c, formula) in formula_snapshot:
        new_c    = col_shift(c, cable_cols)
        new_fml  = shift_formula(formula, cable_cols)
        ws.cell(row=r, column=new_c).value = new_fml

    print(f"  Added {len(cable_cols)} Type column(s)")

wb.save(OUTPUT_FILE)
print(f"\nDone - saved to: {OUTPUT_FILE}")
